from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill #

import requests
import time
import json
import urllib.parse
import os
 

start = 10 # первый элемент таблицы
end = 9999 # последний элемент таблицы

def check_price(view):
    api = 'https://steamcommunity.com/market/priceoverview/?'

    for i in range(start,end):
        
        url = view[f'F{i}'].value
        if url is None:
            break
        else:
            url = url.replace('https://steamcommunity.com/market/listings/730/', '')
            url = url.lstrip().rstrip() #обрезать пробелы в конце/начале строки, чтобы возвращал норм json

            print(f'{urllib.parse.unquote(url)} doing...')
            # параметры запроса
            payload = {'appid':'730','currency':'5', 'market_hash_name':f'{urllib.parse.unquote(url)}'}
            
            # получаем данные
            while True:
                req = requests.get(api, params=payload) # Посылаем запрос к API
                data = req.content.decode() # Декодируем его ответ, чтобы Кириллица отображалась корректно
                if req.status_code == 200:
                    req.close()
                    break
                req.close()
                time.sleep(5)

            #обрабатываем данные
            jsObj = json.loads(data)
            now_price = float(jsObj['lowest_price'].replace('pуб.', '').replace(',','.'))

            view[f'D{i}'].value = now_price

            buy_price = view[f'B{i}'].value 

            # сравниваем цены, чтобы задать фон ячейки
            if now_price - buy_price < 0:
                view[f'E{i}'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101') #красный
            else:
                view[f'E{i}'].fill = PatternFill(fill_type='solid', start_color='00ff00', end_color='00ff00') #зеленый

            print(f'{urllib.parse.unquote(url)} done')
                
            time.sleep(1)
        

def start_prog():
    file = 'data.xlsx'
    # открываем excel файл и получаем лист1
    wb = load_workbook(file)
    view = wb["Лист1"]

    # print_data(view)
    check_price(view)

    sum = 0
    for i in range(start,end):
        if view[f'D{i}'].value is None:
            pass
        else:
            sum += view[f'D{i}'].value * view[f'C{i}'].value

    for i in range(start,end):
        if view[f'K{i}'].value is None:
            view[f'K{i}'].value = sum
            view[f'L{i}'].value = time.ctime()
            break
            
    print(sum)

    wb.save(file)

    os.startfile(file)   


start_prog()


